#!/usr/bin/env python3
"""
Build content_calendar.xlsx for WFIA content tracking.

Three tabs:
  1. Weekly Schedule  — pre-populated through end of Q3 2026 with all recurring slots
  2. Facebook Post Bank — categorized post pipeline
  3. Legend — status definitions and notes

Pulls the Midcard Files priority queue from Postgres so the queue stays in sync
with the source of truth.
"""
import datetime as dt
import sys
from pathlib import Path

import psycopg
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CALENDAR_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = CALENDAR_DIR.parent
APP_DIR = PROJECT_ROOT / "app"
sys.path.insert(0, str(APP_DIR))

from config import CONFIG  # noqa: E402

OUT = CALENDAR_DIR / "content_calendar.xlsx"
# Output: calendar/content_calendar.xlsx

# Calendar window: this week through end of Q3 2026
START = dt.date(2026, 4, 27)   # Monday Apr 27, 2026
END   = dt.date(2026, 9, 28)   # Monday Sept 28, 2026 (~22 weeks)

# Day-of-week assignments (Mon=0, Tue=1, ..., Sun=6)
SLOT_DAY = {
    "Revisiting":          0,  # Monday
    "Midcard Files":       2,  # Wednesday
    "Entrance Music Short":4,  # Friday
    "Longform Feature":    1,  # First Tuesday of month
    "Book Review":         1,  # Third Tuesday of month
}

# ---- Style helpers --------------------------------------------------------
FONT = "Arial"
HDR_FONT = Font(name=FONT, size=11, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", start_color="1F3A5F")
SUBHDR_FONT = Font(name=FONT, size=10, bold=True)
BODY_FONT = Font(name=FONT, size=10)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT_FILL = PatternFill("solid", start_color="F4F6F8")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def style_body(ws, start_row, end_row, ncols, alternate=True):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = LEFT_WRAP
            cell.border = BORDER
            if alternate and (r - start_row) % 2 == 1:
                cell.fill = ALT_FILL


# ---- Pull Midcard queue from Postgres ------------------------------------

def midcard_queue():
    """Return list of (priority, ring_name, why) ordered by priority then name."""
    with psycopg.connect(CONFIG.database_url) as conn:
        cur = conn.execute("""
            SELECT midcard_files_priority, primary_ring_name, why_they_mattered
            FROM wrestlers
            WHERE midcard_files_status = 'queued'
            ORDER BY midcard_files_priority NULLS LAST, primary_ring_name
        """)
        return cur.fetchall()


# ---- Build weekly schedule ----------------------------------------------

def first_weekday_of_month(year, month, weekday):
    d = dt.date(year, month, 1)
    while d.weekday() != weekday:
        d += dt.timedelta(days=1)
    return d


def nth_weekday_of_month(year, month, weekday, n):
    d = first_weekday_of_month(year, month, weekday)
    return d + dt.timedelta(weeks=n - 1)


def weeks_between(start, end):
    """Yield Mondays from start (assumed Monday) through end inclusive."""
    d = start
    while d <= end:
        yield d
        d += dt.timedelta(days=7)


def build_schedule_rows(queue):
    """Return list of dicts ready to write to the schedule sheet."""
    rows = []
    queue_idx = 0
    queue_n = len(queue)

    # First, generate monthly entries (Longform + Book Review)
    monthly = set()
    for year in (2026,):
        for month in range(START.month, END.month + 1):
            longform_d = first_weekday_of_month(year, month, SLOT_DAY["Longform Feature"])
            review_d   = nth_weekday_of_month(year, month, SLOT_DAY["Book Review"], 3)
            if START <= longform_d <= END:
                monthly.add((longform_d, "Longform Feature"))
            if START <= review_d <= END:
                monthly.add((review_d, "Book Review"))

    # Walk weeks
    for monday in weeks_between(START, END):
        # Revisiting (Mon)
        rows.append({
            "week_of": monday,
            "date": monday + dt.timedelta(days=SLOT_DAY["Revisiting"]),
            "type": "Revisiting",
            "title": "",
            "status": "idea",
            "notes": "Re-read older WFIA newsletter article (or other publication w/ permission); fresh take or remix.",
        })
        # Midcard Files (Wed)
        if queue_idx < queue_n:
            prio, name, why = queue[queue_idx]
            queue_idx += 1
            title = f"The Midcard Files: {name}"
            note = (why or "")[:140]
        else:
            title = "The Midcard Files: TBD"
            note = "Queue exhausted — pull next from wrestlers table."
        rows.append({
            "week_of": monday,
            "date": monday + dt.timedelta(days=SLOT_DAY["Midcard Files"]),
            "type": "Midcard Files",
            "title": title,
            "status": "idea",
            "notes": note,
        })
        # Entrance Music Short (Fri)
        rows.append({
            "week_of": monday,
            "date": monday + dt.timedelta(days=SLOT_DAY["Entrance Music Short"]),
            "type": "Entrance Music Short",
            "title": "",
            "status": "idea",
            "notes": "One theme, one story — pulled from longform research doc.",
        })

    # Add monthly rows
    for d, kind in monthly:
        notes = {
            "Longform Feature": "First of month — first feature is full history of entrance music.",
            "Book Review":      "Third Tuesday of month — one wrestling book, taken seriously.",
        }[kind]
        rows.append({
            "week_of": d - dt.timedelta(days=d.weekday()),  # Monday of that week
            "date": d,
            "type": kind,
            "title": "",
            "status": "idea",
            "notes": notes,
        })

    rows.sort(key=lambda r: (r["date"], r["type"]))
    return rows


# ---- Build Facebook post bank --------------------------------------------

FB_BANK_ROWS = [
    # (category, post_hook, target_date, status, source_or_note)
    ("Magazine Archive", "Cover scan + 2-line context — pick a striking cover from the periodical catalog", "", "idea", "Pull from Periodical Catalog Google Sheet."),
    ("Magazine Archive", "Letter page highlight — territorial fan voice from a specific issue",            "", "idea", ""),
    ("Magazine Archive", "Apter-mag rankings page — show how a forgotten worker ranked in a given year",   "", "idea", ""),
    ("Magazine Archive", "Vintage ad — wrestling magazine merch ads as cultural artifact",                  "", "idea", ""),
    ("Midcard Appreciation", "One-photo + 3-line tribute to the Midcard Files subject of the week",         "", "idea", "Auto-syncable to weekly Midcard piece."),
    ("Midcard Appreciation", "Ring-name etymology post — where did the name come from",                     "", "idea", ""),
    ("Midcard Appreciation", "Tag team appreciation — pick a team from the join-table data",                "", "idea", ""),
    ("Bibliography Teaser", "New entry highlight from the wrestling bibliography",                          "", "idea", "Pull from the Postgres bibliography."),
    ("Bibliography Teaser", "Out-of-print book worth tracking down — used market notes",                    "", "idea", ""),
    ("Bibliography Teaser", "Book review preview — one paragraph, full review on Tuesday",                  "", "idea", "Coordinate w/ monthly book review schedule."),
    ("Discussion Prompt", "Which territorial referee was the best in-ring storyteller?",                    "", "idea", ""),
    ("Discussion Prompt", "Underrated tag teams of the 1980s — drop one in the comments",                   "", "idea", ""),
    ("Discussion Prompt", "Best house-show town in the territorial era?",                                   "", "idea", ""),
    ("Discussion Prompt", "Manager you'd most want to see in a modern angle?",                              "", "idea", ""),
    ("Footage Note", "Where to find rare matches — point to a specific YouTube channel or trader",          "", "idea", ""),
    ("Footage Note", "First-time-watching post — short reaction to a recently-surfaced bout",               "", "idea", ""),
]


# ---- Workbook build -------------------------------------------------------

def build():
    wb = Workbook()

    # ---- Tab 1: Weekly Schedule ------------------------------------------
    ws = wb.active
    ws.title = "Weekly Schedule"
    headers = ["Week Of", "Date", "Day", "Content Type", "Working Title", "Status", "Notes"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    queue = midcard_queue()
    rows = build_schedule_rows(queue)

    DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for r in rows:
        ws.append([
            r["week_of"].isoformat(),
            r["date"].isoformat(),
            DAY_NAMES[r["date"].weekday()],
            r["type"],
            r["title"],
            r["status"],
            r["notes"],
        ])
    style_body(ws, 2, ws.max_row, len(headers))
    widths = [12, 12, 6, 22, 42, 12, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---- Tab 2: Facebook Post Bank ---------------------------------------
    fb = wb.create_sheet("FB Post Bank")
    fb_headers = ["Category", "Post Hook", "Target Date", "Status", "Source / Notes"]
    fb.append(fb_headers)
    style_header_row(fb, 1, len(fb_headers))
    for row in FB_BANK_ROWS:
        fb.append(list(row))
    style_body(fb, 2, fb.max_row, len(fb_headers))
    fb_widths = [22, 60, 12, 12, 50]
    for i, w in enumerate(fb_widths, 1):
        fb.column_dimensions[get_column_letter(i)].width = w
    fb.freeze_panes = "A2"

    # ---- Tab 3: Legend ---------------------------------------------------
    lg = wb.create_sheet("Legend")
    lg["A1"] = "Status legend"
    lg["A1"].font = Font(name=FONT, size=12, bold=True)
    legend = [
        ("idea",       "Slot exists in calendar; nothing drafted yet."),
        ("drafting",   "First draft in progress."),
        ("edited",     "Draft is at WFIA-ready quality; awaiting Stephan Watts/Brian Ferguson sign-off."),
        ("scheduled",  "Approved and queued in publishing system."),
        ("published",  "Live."),
        ("declined",   "Pulled from queue — reason in Notes."),
    ]
    lg.append([])
    lg.append(["Status", "Meaning"])
    style_header_row(lg, 3, 2)
    for s, meaning in legend:
        lg.append([s, meaning])
    style_body(lg, 4, lg.max_row, 2)

    lg.append([])
    lg["A" + str(lg.max_row + 1)] = "Recurring slot defaults"
    lg["A" + str(lg.max_row)].font = Font(name=FONT, size=12, bold=True)
    lg.append([])
    lg.append(["Slot", "Day"])
    style_header_row(lg, lg.max_row, 2)
    for slot, day in [
        ("Revisiting", "Monday"),
        ("Midcard Files", "Wednesday"),
        ("Entrance Music Short", "Friday"),
        ("Longform Feature", "1st Tuesday of month"),
        ("Book Review", "3rd Tuesday of month"),
    ]:
        lg.append([slot, day])
    style_body(lg, lg.max_row - 4, lg.max_row, 2)

    lg.append([])
    lg["A" + str(lg.max_row + 1)] = "Source-of-truth pointers"
    lg["A" + str(lg.max_row)].font = Font(name=FONT, size=12, bold=True)
    lg.append([])
    pointers = [
        ("Midcard queue",        "wrestlers table in Postgres (status='queued')"),
        ("Bibliography",         "books / authors tables in Postgres; markdown views in /bibliography/markdown/"),
        ("Periodical catalog",   "Google Sheet 1iPtvpDnv8i2vkB6cSmy_jTCE1GQfAqABO14x-DkndRY"),
        ("Magazine scans",       "Google Drive folder 10AEI9MJac-jbLqbGb9WlF4F_GC0EAGDR"),
        ("Newsletter archive",   "WFIA — pending refreshed password"),
    ]
    lg.append(["What", "Where"])
    style_header_row(lg, lg.max_row, 2)
    for what, where in pointers:
        lg.append([what, where])
    style_body(lg, lg.max_row - len(pointers) + 1, lg.max_row, 2)

    lg.column_dimensions["A"].width = 26
    lg.column_dimensions["B"].width = 80

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"  Weekly Schedule: {ws.max_row - 1} rows")
    print(f"  FB Post Bank:    {fb.max_row - 1} rows")


if __name__ == "__main__":
    build()
